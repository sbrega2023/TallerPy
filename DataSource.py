'''
Created on 2 nov. 2022

@author: sbrega
'''
'''
PROCESA EXCELL
'''
import openpyxl as OPX
from tkinter import filedialog
class DataSource():
    #EXCEL -
    global sheet
    global wb
    global coord
    global archivo
    archivo = 'D:\Movistar_Argentina\Programacion-Desarrollos\Python\Proyectos\TallerPy\PruebaAutoSap.xlsx'
    coord=[]
    #archivo = filedialog.askopenfilename() #Seleccionamos la ruta del archivo a utilizar.
    wb=OPX.load_workbook(archivo) #Acá va la ruta de donde se toma el formulario.
    #sheet=wb['Hoja1']
    sheet = wb.active
    #Variables inicializadas en el Constructor.
    
    def xlData(self):
        data=[]
        dataFinal=[]
        #col=9    #ESTA ES LA CANTIDAD DE COLUMNAS DEL FORMULARIO
        i=0        #ESTE ES EL CONTADOR DE FILAS(ROWS)
        filas=sheet.iter_rows(min_row=1,min_col=1)        #CANTIDAD DE FILAS (generator object)
        #print(filas,type(filas))
        for rows in filas:
            data.append([])
            #print(rows)            #OBTENGO TODAS LAS FILAS    
            for cell in rows:
                #print(cell.value,type(cell.value))        #OBTENGO TODAS LAS CELDAS
                data[i].append(cell.value)
                #print(data)
            i+=1
            dataFinal = filter(any,data)#Funcion Filter arma una lista con los valores de la funcion "any(iterable)" que son "True"
            #print(len(data))
        return list(dataFinal)
    #===========================================================================
    #celda = list(xlData())
    #print(celda, type(celda))
    #print(len(celda))
    
    #===========================================================================
    def writeXl(self, celda, validacion):
        sheet[celda] = validacion
        wb.save(archivo)
        #print(sheet)
    
    def causa(self, celda, causa):
        sheet[celda] = causa
        wb.save(archivo)
        #print(sheet)
           
    def getCoord(self):
        i=0
        for row in sheet.iter_rows():
            coord.append([])
            for cell in row:
                coord[i].append(cell.coordinate)
                #print(cell.coordinate)
            i+=1
        return coord               
        
    #celda = getCoord()
    #print(celda[1][5], type (celda[1][5]))    
    #xlData()    
    #writeXl('F2','procesado')
    
    
        
    def __init__(self):#Constructor
        print("CONSTRUCTOR DataSource")

#causa = DataSource()
#causa.causa('G6', 'No se puede procesar el campo')
        
        
        